# Python 3
# Initial Commit on Aug 8
# Imports
import os
from tkinter import *
from tkinter import ttk
from tkinter.filedialog import *
from tkinter.messagebox import *
import pandas as pd
import openpyxl 

from fetch_sql import *

import sqlite3


def get_bill():
    showinfo('Under development', message='This feature is currently not available.\n Once the feature is ready '
                                          'developer will notify.')


def to_excel():
    new_df = pd.read_sql("SELECT * FROM Database", sqlite3.connect("Employees.db"))
    new_df.to_excel("out.xlsx", index=False)

    no_row = get_row_count()
    work_row = no_row + 2
    basic = get_total_basic()
    day_rate = get_total_rate_day()
    hrs_rate = get_total_rate_ot()
    sun_rate = get_total_rate_sun()
    total_salary = get_total_salary()
    r_off = get_total_r_off()
    net = get_total_net()

    wb = openpyxl.reader.excel.load_workbook(filename='out.xlsx')
    ws = wb.active
    ws.cell(row=work_row, column=1, value='Total')
    ws.cell(row=work_row, column=3, value=basic)
    ws.cell(row=work_row, column=6, value=day_rate)
    ws.cell(row=work_row, column=8, value=hrs_rate)
    ws.cell(row=work_row, column=10, value=sun_rate)
    ws.cell(row=work_row, column=11, value=total_salary)
    ws.cell(row=work_row, column=12, value=r_off)
    ws.cell(row=work_row, column=13, value=net)
    wb.save('out.xlsx')
    wb.close()

    save_file = asksaveasfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    data = pd.read_excel('out.xlsx')
    data.to_excel(save_file + ".xlsx", index=False)
    os.remove('out.xlsx')


def add():
    try:
        db = sqlite3.connect('Employees.db')
        cursor = db.cursor()
        cursor.execute(f"""INSERT INTO  Database VALUES({emp_id_var.get()},
                                                            '{emp_name_var.get()}',
                                                            {basic_salary_var.get()},
                                                            {p_day_var.get()},
                                                            {a_day_var.get()},
                                                            {day_rate_var.get()},
                                                            {ot_hrs_var.get()},
                                                            {hrs_rate_var.get()},
                                                            {extra_sun_var.get()},
                                                            {sun_rate_var.get()},
                                                            {total_salary_var.get()},
                                                            {r_off_var.get()},
                                                            {net_salary_var.get()}
                       )""")
        db.commit()
        db.close()  
        clear()
        fetch_data()
    except sqlite3.IntegrityError:
        showerror('Id Already Exists ', message=f'Check the Employee Id. \n Employee Id should be unique. \n'
                                                f'Employee Id: {emp_id_var.get()} belongs to some one else.')

    except TclError:
        showerror('Insert Data', message='Insert data first. \n Fill in the blank entries with zero \n for the first '
                                         'time.')


def delete():
    try:
        db = sqlite3.connect('Employees.db')
        cursor = db.cursor()
        cursor.execute(f"""DELETE FROM Database WHERE emp_id = {emp_id_var.get()}""")
        db.commit()
        db.close()
        clear()
        fetch_data()
    except TclError:
        showerror("Delete Error", message='Select existing entry first. \n Click on the entry to be deleted')


def update():
    try:
        db = sqlite3.connect('Employees.db')
        cursor = db.cursor()

        cursor.execute(f"""UPDATE Database SET emp_name = '{emp_name_var.get()}', 'basic'= {basic_salary_var.get()},
                                                     p_day= {p_day_var.get()},
                                                     a_day= {a_day_var.get()},
                                                     day_rate = {day_rate_var.get()},
                                                     ot_hrs = {ot_hrs_var.get()},
                                                     ot_rate = {hrs_rate_var.get()},
                                                     ex_sun = {extra_sun_var.get()},
                                                     sun_rate = {sun_rate_var.get()},
                                                     total_salary = {total_salary_var.get()},
                                                     r_off = {r_off_var.get()},
                                                     net_salary = {net_salary_var.get()}
                                                    WHERE emp_id = {emp_id_var.get()}
        """)

        db.commit()
        db.close()
        fetch_data()
    except TclError:
        showerror("Update Error", message='Select existing entry first. \n Click on the entry to be updated')


def clear():
    emp_id_var.set("")
    basic_salary_var.set("")
    emp_name_var.set("")

    p_day_var.set("")
    a_day_var.set("")
    day_rate_var.set("")

    ot_hrs_var.set("")
    hrs_rate_var.set("")

    extra_sun_var.set("")
    sun_rate_var.set("")

    total_salary_var.set("")
    r_off_var.set("")
    net_salary_var.set("")


def cal_salary():
    try:
        basic = int(basic_salary_var.get())
        days = no_days_var.get()
        days = days - 4 # 4-5 sundays
        day_rate_var.set(f"{basic / days}")
        day_rate = basic / days

        hrs_rate_var.set(f"{day_rate / 8}")
        hrs_rate = day_rate / 8

        sun_rate_var.set(f"{day_rate}")
        sun_rate = day_rate

        p = p_day_var.get() #present no of days
        # a = a_day_var.get()
        a = days - p #absentr no of days
        a_day_var.set(f'{a}')
        h = ot_hrs_var.get() #overtime hours
        s = extra_sun_var.get() #overtime sundays

        total = (p * day_rate) + (h * hrs_rate) + (s * sun_rate)

        total_salary_var.set(f"{total}")
        r = round(total)
        r_off_var.set(f"{r}")
        net_salary_var.set(f"{r}")
        update()
    except TclError:
        showerror("Calculate Error", message='Select existing entry first. \n Click on the entry to be Calculate. \n'
                                             'Tip: Check the No. of Days entry on the top right.')


def fetch_data():
    db = sqlite3.connect('Employees.db')
    cursor = db.cursor()
    cursor.execute("""SELECT * FROM Database""")
    rows = cursor.fetchall()
    if len(rows) != 0:
        emp_table.delete(*emp_table.get_children())
        for row in rows:
            emp_table.insert('', END, values=row)

    db.commit()
    db.close()


def get_data(a):
    try:
        cursor_row = emp_table.focus()
        content = emp_table.item(cursor_row)
        row = content['values']
        emp_id_var.set(row[0])
        emp_name_var.set(row[1])
        basic_salary_var.set(row[2])

        p_day_var.set(row[3])
        a_day_var.set(row[4])
        day_rate_var.set(row[5])

        ot_hrs_var.set(row[6])
        hrs_rate_var.set(row[7])

        extra_sun_var.set(row[8])
        sun_rate_var.set(row[9])

        total_salary_var.set(row[10])
        r_off_var.set(row[11])
        net_salary_var.set(row[12])
    except IndexError:
        print(a)


def search():
    if search_by_var.get() == '':
        showerror("Search Error", message='Select an option to search based on the option.')
    elif search_text_var.get() == '':
        showerror("Search Error", message='Search field is empty.')
    else:
        global emp, e_id, msg
        db = sqlite3.connect('Employees.db')
        cursor = db.cursor()
        if str(search_by_var.get()) == 'Employee Id':
            e_id = int(search_text_var.get())
            emp = 'emp_id'
            msg = 'id'
            cursor.execute(f"""SELECT * FROM Database WHERE emp_id = {e_id}""")
        elif str(search_by_var.get()) == 'Name':
            e_name = str(search_text_var.get())
            msg = 'name'
            cursor.execute(f"""SELECT * FROM Database WHERE emp_name LIKE '%{e_name}%'""")
        rows = cursor.fetchall()
        if len(rows) != 0:
            emp_table.delete(*emp_table.get_children())
            for row in rows:
                emp_table.insert('', END, values=row)
        else:
            showerror("Search Error", message=f"Couldn't find Employee with {msg}: {search_text_var.get()}")

        db.commit()
        db.close()


def main():
    """This class configures and populates the toplevel window.
               top is the toplevel containing window."""
    _bgcolor = '#d9d9d9'  # X11 color: 'gray85'
    _fgcolor = '#000000'  # X11 color: 'black'
    _compcolor = '#d9d9d9'  # X11 color: 'gray85'
    _ana1color = '#d9d9d9'  # X11 color: 'gray85'
    _ana2color = '#ececec'  # Closest X11 color: 'gray92'
    # font13 = "-family {Segoe UI} -size 9"

    root = Tk()
    root.state('zoomed')
    root.geometry("1920x1057")
    root.minsize(120, 1)
    root.maxsize(1924, 1061)
    root.resizable(1, 1)
    root.title("Salary Management - Employee Database")
    root.configure(background="#80ffff")
    root.configure(highlightbackground="#d9d9d9")
    root.configure(highlightcolor="black")

    menubar = Menu(root, font="-family {Arial} -size 9", bg=_bgcolor, fg=_fgcolor)
    file_menu = Menu(menubar, tearoff=0)
    file_menu.add_command(label="Generate Excel", command=to_excel)

    file_menu.add_separator()

    file_menu.add_command(label="Generate Bill")

    menubar.add_cascade(label="Gen Excel", menu=file_menu)
    root.configure(menu=menubar)

    # All Variables
    global no_days_var, emp_id_var, basic_salary_var, p_day_var, a_day_var, day_rate_var, ot_hrs_var
    global extra_sun_var, sun_rate_var, total_salary_var, r_off_var, net_salary_var, hrs_rate_var
    global emp_table, emp_name_var, search_by_var, search_text_var

    no_days_var = IntVar()
    emp_id_var = IntVar()
    basic_salary_var = IntVar()
    emp_name_var = StringVar()

    p_day_var = IntVar()
    a_day_var = IntVar()
    day_rate_var = DoubleVar()

    ot_hrs_var = IntVar()
    hrs_rate_var = DoubleVar()

    extra_sun_var = IntVar()
    sun_rate_var = DoubleVar()

    total_salary_var = DoubleVar()
    r_off_var = DoubleVar()
    net_salary_var = DoubleVar()

    search_by_var = StringVar()
    search_text_var = StringVar()

    Head = Label(root)
    Head.place(relx=0.0, rely=0.0, height=70, width=1920)
    Head.configure(activebackground="#f9f9f9")
    Head.configure(activeforeground="black")
    Head.configure(background="#80ffff")
    Head.configure(disabledforeground="#a3a3a3")
    Head.configure(font="-family {Cambria Math} -size 24 -weight bold")
    Head.configure(foreground="#0000ff")
    Head.configure(highlightbackground="#d9d9d9")
    Head.configure(highlightcolor="black")
    Head.configure(text='''Employee Database''')

    frame_detail = Frame(root)
    frame_detail.place(relx=0.005, rely=0.076, relheight=0.921, relwidth=0.328)
    frame_detail.configure(relief='ridge')
    frame_detail.configure(borderwidth="5")
    frame_detail.configure(relief="ridge")
    frame_detail.configure(background="#80ffff")
    frame_detail.configure(highlightbackground="#d9d9d9")
    frame_detail.configure(highlightcolor="black")

    head_detail_l = Label(frame_detail)
    head_detail_l.place(relx=0, rely=0, height=70, width=619)
    head_detail_l.configure(activebackground="#f9f9f9")
    head_detail_l.configure(activeforeground="black")
    head_detail_l.configure(background="#b1d3e3")
    head_detail_l.configure(disabledforeground="#a3a3a3")
    head_detail_l.configure(font="-family {Arial} -size 24")
    head_detail_l.configure(foreground="#000000")
    head_detail_l.configure(highlightbackground="#d9d9d9")
    head_detail_l.configure(highlightcolor="black")
    head_detail_l.configure(text='''Employee Details''')

    frame_det_btn = Frame(frame_detail)
    frame_det_btn.place(relx=0.008, rely=0.855, relheight=0.139, relwidth=0.984)
    frame_det_btn.configure(relief='ridge')
    frame_det_btn.configure(borderwidth="6")
    frame_det_btn.configure(relief="ridge")
    frame_det_btn.configure(background="#b1d3e3")
    frame_det_btn.configure(highlightbackground="#d9d9d9")
    frame_det_btn.configure(highlightcolor="black")

    btn_add = Button(frame_det_btn, command=add)
    btn_add.place(relx=0.018, rely=0.081, height=49, width=130)
    btn_add.configure(activebackground="#ececec")
    btn_add.configure(activeforeground="#000000")
    btn_add.configure(background="#80ffff")
    btn_add.configure(disabledforeground="#a3a3a3")
    btn_add.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_add.configure(foreground="#000000")
    btn_add.configure(highlightbackground="#d9d9d9")
    btn_add.configure(highlightcolor="black")
    btn_add.configure(pady="0")
    btn_add.configure(text='''Add''')

    btn_del = Button(frame_det_btn, command=delete)
    btn_del.place(relx=0.263, rely=0.081, height=49, width=130)
    btn_del.configure(activebackground="#ececec")
    btn_del.configure(activeforeground="#000000")
    btn_del.configure(background="#80ffff")
    btn_del.configure(disabledforeground="#a3a3a3")
    btn_del.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_del.configure(foreground="#000000")
    btn_del.configure(highlightbackground="#d9d9d9")
    btn_del.configure(highlightcolor="black")
    btn_del.configure(pady="0")
    btn_del.configure(text='''Delete''')

    btn_upd = Button(frame_det_btn, command=update)
    btn_upd.place(relx=0.509, rely=0.081, height=49, width=130)
    btn_upd.configure(activebackground="#ececec")
    btn_upd.configure(activeforeground="#000000")
    btn_upd.configure(background="#80ffff")
    btn_upd.configure(disabledforeground="#a3a3a3")
    btn_upd.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_upd.configure(foreground="#000000")
    btn_upd.configure(highlightbackground="#d9d9d9")
    btn_upd.configure(highlightcolor="black")
    btn_upd.configure(pady="0")
    btn_upd.configure(text='''Update''')

    btn_clear = Button(frame_det_btn, command=clear)
    btn_clear.place(relx=0.754, rely=0.081, height=49, width=130)
    btn_clear.configure(activebackground="#ececec")
    btn_clear.configure(activeforeground="#000000")
    btn_clear.configure(background="#80ffff")
    btn_clear.configure(disabledforeground="#a3a3a3")
    btn_clear.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_clear.configure(foreground="#000000")
    btn_clear.configure(highlightbackground="#d9d9d9")
    btn_clear.configure(highlightcolor="black")
    btn_clear.configure(pady="0")
    btn_clear.configure(text='''Clear''')

    btn_cal = Button(frame_det_btn, command=cal_salary)
    btn_cal.place(relx=0.263, rely=0.526, height=49, width=282)
    btn_cal.configure(activebackground="#ececec")
    btn_cal.configure(activeforeground="#000000")
    btn_cal.configure(background="#80ffff")
    btn_cal.configure(disabledforeground="#a3a3a3")
    btn_cal.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_cal.configure(foreground="#000000")
    btn_cal.configure(highlightbackground="#d9d9d9")
    btn_cal.configure(highlightcolor="black")
    btn_cal.configure(pady="0")
    btn_cal.configure(text='''Calculate Salary''')

    emp_id_l = Label(frame_detail)
    emp_id_l.place(relx=0.048, rely=0.113, height=40, width=170)
    emp_id_l.configure(activebackground="#80ffff")
    emp_id_l.configure(activeforeground="black")
    emp_id_l.configure(background="#80ffff")
    emp_id_l.configure(disabledforeground="#a3a3a3")
    emp_id_l.configure(font="-family {Arial} -size 14")
    emp_id_l.configure(foreground="black")
    emp_id_l.configure(highlightbackground="#80ffff")
    emp_id_l.configure(highlightcolor="black")
    emp_id_l.configure(text='''Employee Id''')

    emp_id_e = Entry(frame_detail, textvariable=emp_id_var)
    emp_id_e.place(relx=0.318, rely=0.113, height=40, relwidth=0.102)
    emp_id_e.configure(background="white")
    emp_id_e.configure(disabledforeground="#a3a3a3")
    emp_id_e.configure(font="-family {Courier New} -size 10")
    emp_id_e.configure(foreground="black")
    emp_id_e.configure(highlightbackground="#d9d9d9")
    emp_id_e.configure(highlightcolor="black")
    emp_id_e.configure(insertbackground="black")
    emp_id_e.configure(selectbackground="blue")
    emp_id_e.configure(selectforeground="white")

    emp_name_l = Label(frame_detail)
    emp_name_l.place(relx=0.048, rely=0.185, height=40, width=170)
    emp_name_l.configure(activebackground="#f9f9f9")
    emp_name_l.configure(activeforeground="black")
    emp_name_l.configure(background="#80ffff")
    emp_name_l.configure(disabledforeground="#a3a3a3")
    emp_name_l.configure(font="-family {Arial} -size 14")
    emp_name_l.configure(foreground="black")
    emp_name_l.configure(highlightbackground="#d9d9d9")
    emp_name_l.configure(highlightcolor="black")
    emp_name_l.configure(text='''Name''')

    basic_l = Label(frame_detail)
    basic_l.place(relx=0.048, rely=0.257, height=40, width=170)
    basic_l.configure(activebackground="#f9f9f9")
    basic_l.configure(activeforeground="black")
    basic_l.configure(background="#80ffff")
    basic_l.configure(disabledforeground="#a3a3a3")
    basic_l.configure(font="-family {Arial} -size 14")
    basic_l.configure(foreground="black")
    basic_l.configure(highlightbackground="#d9d9d9")
    basic_l.configure(highlightcolor="black")
    basic_l.configure(text='''Basic Salary''')

    pd_l = Label(frame_detail)
    pd_l.place(relx=0.048, rely=0.329, height=40, width=170)
    pd_l.configure(activebackground="#f9f9f9")
    pd_l.configure(activeforeground="black")
    pd_l.configure(background="#80ffff")
    pd_l.configure(disabledforeground="#a3a3a3")
    pd_l.configure(font="-family {Arial} -size 14")
    pd_l.configure(foreground="black")
    pd_l.configure(highlightbackground="#d9d9d9")
    pd_l.configure(highlightcolor="black")
    pd_l.configure(text='''Present Days''')

    abd_l = Label(frame_detail)
    abd_l.place(relx=0.334, rely=0.329, height=40, width=170)
    abd_l.configure(activebackground="#f9f9f9")
    abd_l.configure(activeforeground="black")
    abd_l.configure(background="#80ffff")
    abd_l.configure(disabledforeground="#a3a3a3")
    abd_l.configure(font="-family {Arial} -size 14")
    abd_l.configure(foreground="black")
    abd_l.configure(highlightbackground="#d9d9d9")
    abd_l.configure(highlightcolor="black")
    abd_l.configure(text='''Absent Days''')

    day_rate_l = Label(frame_detail)
    day_rate_l.place(relx=0.62, rely=0.329, height=40, width=170)
    day_rate_l.configure(activebackground="#f9f9f9")
    day_rate_l.configure(activeforeground="black")
    day_rate_l.configure(background="#80ffff")
    day_rate_l.configure(disabledforeground="#a3a3a3")
    day_rate_l.configure(font="-family {Arial} -size 14")
    day_rate_l.configure(foreground="black")
    day_rate_l.configure(highlightbackground="#d9d9d9")
    day_rate_l.configure(highlightcolor="black")
    day_rate_l.configure(text='''Daily Rate''')

    ot_l = Label(frame_detail)
    ot_l.place(relx=0.048, rely=0.472, height=40, width=170)
    ot_l.configure(activebackground="#f9f9f9")
    ot_l.configure(activeforeground="black")
    ot_l.configure(background="#80ffff")
    ot_l.configure(disabledforeground="#a3a3a3")
    ot_l.configure(font="-family {Arial} -size 14")
    ot_l.configure(foreground="black")
    ot_l.configure(highlightbackground="#d9d9d9")
    ot_l.configure(highlightcolor="black")
    ot_l.configure(text='''Over Time (Hrs)''')

    hrs_rate_l = Label(frame_detail)
    hrs_rate_l.place(relx=0.477, rely=0.472, height=40, width=170)
    hrs_rate_l.configure(activebackground="#f9f9f9")
    hrs_rate_l.configure(activeforeground="black")
    hrs_rate_l.configure(background="#80ffff")
    hrs_rate_l.configure(disabledforeground="#a3a3a3")
    hrs_rate_l.configure(font="-family {Arial} -size 14")
    hrs_rate_l.configure(foreground="black")
    hrs_rate_l.configure(highlightbackground="#d9d9d9")
    hrs_rate_l.configure(highlightcolor="black")
    hrs_rate_l.configure(text='''Hourly Rate''')

    sun_l = Label(frame_detail)
    sun_l.place(relx=0.048, rely=0.544, height=40, width=170)
    sun_l.configure(activebackground="#f9f9f9")
    sun_l.configure(activeforeground="black")
    sun_l.configure(background="#80ffff")
    sun_l.configure(disabledforeground="#a3a3a3")
    sun_l.configure(font="-family {Arial} -size 14")
    sun_l.configure(foreground="black")
    sun_l.configure(highlightbackground="#d9d9d9")
    sun_l.configure(highlightcolor="black")
    sun_l.configure(text='''Extra Sunday (0-5)''')

    sun_rate_l = Label(frame_detail)
    sun_rate_l.place(relx=0.477, rely=0.554, height=40, width=170)
    sun_rate_l.configure(activebackground="#f9f9f9")
    sun_rate_l.configure(activeforeground="black")
    sun_rate_l.configure(background="#80ffff")
    sun_rate_l.configure(disabledforeground="#a3a3a3")
    sun_rate_l.configure(font="-family {Arial} -size 14")
    sun_rate_l.configure(foreground="black")
    sun_rate_l.configure(highlightbackground="#d9d9d9")
    sun_rate_l.configure(highlightcolor="black")
    sun_rate_l.configure(text='''Extra Sun Rate''')

    t_salary_l = Label(frame_detail)
    t_salary_l.place(relx=0.153, rely=0.616, height=40, width=170)
    t_salary_l.configure(activebackground="#f9f9f9")
    t_salary_l.configure(activeforeground="black")
    t_salary_l.configure(background="#80ffff")
    t_salary_l.configure(disabledforeground="#a3a3a3")
    t_salary_l.configure(font="-family {Arial} -size 14")
    t_salary_l.configure(foreground="black")
    t_salary_l.configure(highlightbackground="#d9d9d9")
    t_salary_l.configure(highlightcolor="black")
    t_salary_l.configure(text='''Total Salary''')

    r_off_l = Label(frame_detail)
    r_off_l.place(relx=0.576, rely=0.616, height=40, width=170)
    r_off_l.configure(activebackground="#f9f9f9")
    r_off_l.configure(activeforeground="black")
    r_off_l.configure(background="#80ffff")
    r_off_l.configure(disabledforeground="#a3a3a3")
    r_off_l.configure(font="-family {Arial} -size 14")
    r_off_l.configure(foreground="black")
    r_off_l.configure(highlightbackground="#d9d9d9")
    r_off_l.configure(highlightcolor="black")
    r_off_l.configure(text='''Round Off''')

    net_salary_l = Label(frame_detail)
    net_salary_l.place(relx=0.477, rely=0.76, height=40, width=170)
    net_salary_l.configure(activebackground="#f9f9f9")
    net_salary_l.configure(activeforeground="black")
    net_salary_l.configure(background="#80ffff")
    net_salary_l.configure(disabledforeground="#a3a3a3")
    net_salary_l.configure(font="-family {Arial} -size 14 -weight bold")
    net_salary_l.configure(foreground="black")
    net_salary_l.configure(highlightbackground="#d9d9d9")
    net_salary_l.configure(highlightcolor="black")
    net_salary_l.configure(text='''Net Salary''')

    emp_name_e = Entry(frame_detail, textvariable=emp_name_var)
    emp_name_e.place(relx=0.318, rely=0.185, height=40, relwidth=0.35)
    emp_name_e.configure(background="white")
    emp_name_e.configure(disabledforeground="#a3a3a3")
    emp_name_e.configure(font="-family {Courier New} -size 10")
    emp_name_e.configure(foreground="black")
    emp_name_e.configure(highlightbackground="#d9d9d9")
    emp_name_e.configure(highlightcolor="black")
    emp_name_e.configure(insertbackground="black")
    emp_name_e.configure(selectbackground="blue")
    emp_name_e.configure(selectforeground="white")

    basic_e = Entry(frame_detail, textvariable=basic_salary_var)
    basic_e.place(relx=0.318, rely=0.257, height=40, relwidth=0.159)
    basic_e.configure(background="white")
    basic_e.configure(disabledforeground="#a3a3a3")
    basic_e.configure(font="-family {Courier New} -size 10")
    basic_e.configure(foreground="black")
    basic_e.configure(highlightbackground="#d9d9d9")
    basic_e.configure(highlightcolor="black")
    basic_e.configure(insertbackground="black")
    basic_e.configure(selectbackground="blue")
    basic_e.configure(selectforeground="white")

    pd_e = Entry(frame_detail, textvariable=p_day_var)
    pd_e.place(relx=0.132, rely=0.4, height=40, relwidth=0.102)
    pd_e.configure(background="white")
    pd_e.configure(disabledforeground="#a3a3a3")
    pd_e.configure(font="-family {Courier New} -size 10")
    pd_e.configure(foreground="black")
    pd_e.configure(highlightbackground="#d9d9d9")
    pd_e.configure(highlightcolor="black")
    pd_e.configure(insertbackground="black")
    pd_e.configure(selectbackground="blue")
    pd_e.configure(selectforeground="white")

    abd_e = Entry(frame_detail, textvariable=a_day_var)
    abd_e.place(relx=0.418, rely=0.4, height=40, relwidth=0.102)
    abd_e.configure(background="white")
    abd_e.configure(disabledforeground="#a3a3a3")
    abd_e.configure(font="-family {Courier New} -size 10")
    abd_e.configure(foreground="black")
    abd_e.configure(highlightbackground="#d9d9d9")
    abd_e.configure(highlightcolor="black")
    abd_e.configure(insertbackground="black")
    abd_e.configure(selectbackground="blue")
    abd_e.configure(selectforeground="white")

    day_rate_e = Entry(frame_detail, textvariable=day_rate_var)
    day_rate_e.place(relx=0.676, rely=0.4, height=40, relwidth=0.159)
    day_rate_e.configure(background="white")
    day_rate_e.configure(disabledforeground="#a3a3a3")
    day_rate_e.configure(font="-family {Courier New} -size 10")
    day_rate_e.configure(foreground="black")
    day_rate_e.configure(highlightbackground="#d9d9d9")
    day_rate_e.configure(highlightcolor="black")
    day_rate_e.configure(insertbackground="black")
    day_rate_e.configure(selectbackground="blue")
    day_rate_e.configure(selectforeground="white")

    ot_e = Entry(frame_detail, textvariable=ot_hrs_var)
    ot_e.place(relx=0.318, rely=0.472, height=40, relwidth=0.102)
    ot_e.configure(background="white")
    ot_e.configure(disabledforeground="#a3a3a3")
    ot_e.configure(font="-family {Courier New} -size 10")
    ot_e.configure(foreground="black")
    ot_e.configure(highlightbackground="#d9d9d9")
    ot_e.configure(highlightcolor="black")
    ot_e.configure(insertbackground="black")
    ot_e.configure(selectbackground="blue")
    ot_e.configure(selectforeground="white")

    hrs_rate_e = Entry(frame_detail, textvariable=hrs_rate_var)
    hrs_rate_e.place(relx=0.779, rely=0.472, height=40, relwidth=0.159)
    hrs_rate_e.configure(background="white")
    hrs_rate_e.configure(disabledforeground="#a3a3a3")
    hrs_rate_e.configure(font="-family {Courier New} -size 10")
    hrs_rate_e.configure(foreground="black")
    hrs_rate_e.configure(highlightbackground="#d9d9d9")
    hrs_rate_e.configure(highlightcolor="black")
    hrs_rate_e.configure(insertbackground="black")
    hrs_rate_e.configure(selectbackground="blue")
    hrs_rate_e.configure(selectforeground="white")

    sun_e = Entry(frame_detail, textvariable=extra_sun_var)
    sun_e.place(relx=0.318, rely=0.544, height=40, relwidth=0.102)
    sun_e.configure(background="white")
    sun_e.configure(disabledforeground="#a3a3a3")
    sun_e.configure(font="-family {Courier New} -size 10")
    sun_e.configure(foreground="black")
    sun_e.configure(highlightbackground="#d9d9d9")
    sun_e.configure(highlightcolor="black")
    sun_e.configure(insertbackground="black")
    sun_e.configure(selectbackground="blue")
    sun_e.configure(selectforeground="white")

    sun_rate_e = Entry(frame_detail, textvariable=sun_rate_var)
    sun_rate_e.place(relx=0.779, rely=0.544, height=40, relwidth=0.159)
    sun_rate_e.configure(background="white")
    sun_rate_e.configure(disabledforeground="#a3a3a3")
    sun_rate_e.configure(font="-family {Courier New} -size 10")
    sun_rate_e.configure(foreground="black")
    sun_rate_e.configure(highlightbackground="#d9d9d9")
    sun_rate_e.configure(highlightcolor="black")
    sun_rate_e.configure(insertbackground="black")
    sun_rate_e.configure(selectbackground="blue")
    sun_rate_e.configure(selectforeground="white")

    t_salary_e = Entry(frame_detail, textvariable=total_salary_var)
    t_salary_e.place(relx=0.208, rely=0.688, height=40, relwidth=0.159)
    t_salary_e.configure(background="white")
    t_salary_e.configure(disabledforeground="#a3a3a3")
    t_salary_e.configure(font="-family {Courier New} -size 10")
    t_salary_e.configure(foreground="black")
    t_salary_e.configure(highlightbackground="#d9d9d9")
    t_salary_e.configure(highlightcolor="black")
    t_salary_e.configure(insertbackground="black")
    t_salary_e.configure(selectbackground="blue")
    t_salary_e.configure(selectforeground="white")

    r_off_e = Entry(frame_detail, textvariable=r_off_var)
    r_off_e.place(relx=0.631, rely=0.688, height=40, relwidth=0.159)
    r_off_e.configure(background="white")
    r_off_e.configure(disabledforeground="#a3a3a3")
    r_off_e.configure(font="-family {Courier New} -size 10")
    r_off_e.configure(foreground="black")
    r_off_e.configure(highlightbackground="#d9d9d9")
    r_off_e.configure(highlightcolor="black")
    r_off_e.configure(insertbackground="black")
    r_off_e.configure(selectbackground="blue")
    r_off_e.configure(selectforeground="white")

    net_salary_e = Entry(frame_detail, textvariable=net_salary_var)
    net_salary_e.place(relx=0.779, rely=0.76, height=40, relwidth=0.159)

    net_salary_e.configure(background="white")
    net_salary_e.configure(disabledforeground="#a3a3a3")
    net_salary_e.configure(font="-family {Courier New} -size 10")
    net_salary_e.configure(foreground="black")
    net_salary_e.configure(highlightbackground="#d9d9d9")
    net_salary_e.configure(highlightcolor="black")
    net_salary_e.configure(insertbackground="black")
    net_salary_e.configure(selectbackground="blue")
    net_salary_e.configure(selectforeground="white")

    frame_data = Frame(root)
    frame_data.place(relx=0.344, rely=0.076, relheight=0.921, relwidth=0.651)
    frame_data.configure(relief='ridge')
    frame_data.configure(borderwidth="5")
    frame_data.configure(relief="ridge")
    frame_data.configure(background="#80ffff")
    frame_data.configure(highlightbackground="#d9d9d9")
    frame_data.configure(highlightcolor="black")

    label_search = Label(frame_data)
    label_search.place(relx=0.024, rely=0.005, height=70, width=200)
    label_search.configure(activebackground="#f9f9f9")
    label_search.configure(activeforeground="black")
    label_search.configure(background="#80ffff")
    label_search.configure(disabledforeground="#a3a3a3")
    label_search.configure(font="-family {Arial} -size 24")
    label_search.configure(foreground="#000000")
    label_search.configure(highlightbackground="#d9d9d9")
    label_search.configure(highlightcolor="black")
    label_search.configure(text='''Search By''')

    cmb_search = ttk.Combobox(frame_data, textvariable=search_by_var)
    cmb_search.place(relx=0.2, rely=0.031, relheight=0.024, relwidth=0.16)
    cmb_search['values'] = ('Employee Id', 'Name')
    cmb_search.configure(takefocus="")

    search_e = Entry(frame_data, textvariable=search_text_var)
    search_e.place(relx=0.4, rely=0.031, height=22, relwidth=0.16)
    search_e.configure(background="white")
    search_e.configure(disabledforeground="#a3a3a3")
    search_e.configure(font="-family {Courier New} -size 10")
    search_e.configure(foreground="#000000")
    search_e.configure(highlightbackground="#d9d9d9")
    search_e.configure(highlightcolor="black")
    search_e.configure(insertbackground="black")
    search_e.configure(selectbackground="blue")
    search_e.configure(selectforeground="white")

    btn_search = Button(frame_data, command=search)
    btn_search.place(relx=0.584, rely=0.021, height=49, width=130)
    btn_search.configure(activebackground="#ececec")
    btn_search.configure(activeforeground="#000000")
    btn_search.configure(background="#b1d3e3")
    btn_search.configure(disabledforeground="#a3a3a3")
    btn_search.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_search.configure(foreground="#000000")
    btn_search.configure(highlightbackground="#d9d9d9")
    btn_search.configure(highlightcolor="black")
    btn_search.configure(pady="0")
    btn_search.configure(text='''Search''')

    btn_show = Button(frame_data, command=fetch_data)
    btn_show.place(relx=0.704, rely=0.021, height=49, width=130)
    btn_show.configure(activebackground="#ececec")
    btn_show.configure(activeforeground="#000000")
    btn_show.configure(background="#b1d3e3")
    btn_show.configure(disabledforeground="#a3a3a3")
    btn_show.configure(font="-family {Arial Narrow Special G1} -size 17")
    btn_show.configure(foreground="#000000")
    btn_show.configure(highlightbackground="#d9d9d9")
    btn_show.configure(highlightcolor="black")
    btn_show.configure(pady="0")
    btn_show.configure(text='''Show All''')

    no_day_l = Label(frame_data)
    no_day_l.place(relx=0.824, rely=0.062, height=39, width=140)
    no_day_l.configure(activebackground="#80ffff")
    no_day_l.configure(background="#80ffff")
    no_day_l.configure(disabledforeground="#a3a3a3")
    no_day_l.configure(font="-family {Arial} -size 19")
    no_day_l.configure(foreground="#000000")
    no_day_l.configure(text='''No. of Days''')

    no_day_e = ttk.Combobox(frame_data, textvariable=no_days_var)
    no_day_e.place(relx=0.944, rely=0.062, height=40, relwidth=0.035)
    no_day_e.configure(background="white")
    no_day_e.configure(font="-family {Courier New} -size 10")
    no_day_e.configure(foreground="#000000")
    no_day_e['values'] = (28, 29, 30, 31)

    frame_Table = Frame(frame_data)
    frame_Table.place(relx=0.008, rely=0.113, relheight=0.881, relwidth=0.984)
    frame_Table.configure(relief='groove')
    frame_Table.configure(borderwidth="2")
    frame_Table.configure(relief="groove")
    frame_Table.configure(background="#80ffff")
    frame_Table.configure(highlightbackground="#d9d9d9")
    frame_Table.configure(highlightcolor="black")

    scroll_x = Scrollbar(frame_Table, orient=HORIZONTAL)
    scroll_y = Scrollbar(frame_Table, orient=VERTICAL)

    emp_table = ttk.Treeview(frame_Table, column=("emp_id", 'emp_name', 'basic_salary', 'p_day', 'a_day', 'day_rate',
                                                  'ot_hrs', 'hrs_rate', 'extra_sun', 'sun_rate', 'Total_rate', 'r_off',
                                                  'net_salary'), xscrollcommand=scroll_x.set,
                             yscrollcommand=scroll_y.set)

    scroll_x.pack(side=BOTTOM, fill=X)
    scroll_x.configure(command=emp_table.xview)
    scroll_y.pack(side=RIGHT, fill=Y)
    scroll_y.configure(command=emp_table.yview)

    emp_table.heading("emp_id", text="Employee No.")
    emp_table.heading("emp_name", text="Employee Name")
    emp_table.heading("basic_salary", text="Basic Salary")
    emp_table.heading("p_day", text="Present Days")
    emp_table.heading("a_day", text="Absent Days")
    emp_table.heading("day_rate", text="Daily Rate")
    emp_table.heading("ot_hrs", text="Over Time")
    emp_table.heading("hrs_rate", text="Hourly Rate")
    emp_table.heading("extra_sun", text="Extra Sunday")
    emp_table.heading("sun_rate", text="Sunday Rate")
    emp_table.heading("Total_rate", text="Total Rate")
    emp_table.heading("r_off", text="Round Off")
    emp_table.heading("net_salary", text="Net Salary")

    emp_table['show'] = 'headings'
    emp_table.column("emp_id", width=100)
    emp_table.column("emp_name", width=100)
    emp_table.column("basic_salary", width=100)
    emp_table.column("p_day", width=100)
    emp_table.column("a_day", width=100)
    emp_table.column("day_rate", width=100)
    emp_table.column("ot_hrs", width=100)
    emp_table.column("hrs_rate", width=100)
    emp_table.column("extra_sun", width=100)
    emp_table.column("sun_rate", width=100)
    emp_table.column("Total_rate", width=100)
    emp_table.column("r_off", width=100)
    emp_table.column("net_salary", width=100)

    emp_table.pack(fill=BOTH, expand=1)
    emp_table.bind("<ButtonRelease-1>", get_data)
    fetch_data()
    clear()
    no_days_var.set('30')

    root.mainloop()


if __name__ == '__main__':
    main()
